#!/usr/bin/env python3
"""Die Auszahlungsberichte der Portale (B11).

**Die Frage, aus der das Paket entstand:** „Woher soll ich denn jetzt wissen,
welche Rechnung zu welchem Auszahlungsbetrag gehört?"

Aus Bank und Rechnungen allein ist das nicht zu beantworten. Booking zahlt
wöchentlich gesammelt und netto aus; auf dem Kontoauszug steht nur

    ZAHLUNGSEINGANG NO.vQmQNcef5aDAINyZ/ID.15049295

Kein Gastname, keine Reservierungsnummer, kein Betrag, der zu einer Rechnung
passte – die Provision ist schon abgezogen. Von 65 Auszahlungen ließ sich genau
eine über den Betrag einer einzelnen Rechnung zuordnen.

**Der Bericht schließt die Lücke.** In seiner Spalte „Statement Descriptor"
steht genau das Kürzel aus dem Verwendungszweck, und darunter stehen die
Reservierungen, die in dieser Auszahlung stecken – mit Nummer, Brutto und
Provision. Über die Reservierungsnummer hängt die Smoobu-Buchung, an ihr die
Rechnung. Damit ist die Kette vollständig:

    Bankbewegung → NO.<Kürzel> → Auszahlung → Reservierungsnummer
                 → Smoobu-Buchung → Rechnung

An den echten Daten (Januar bis Juli 2026) gemessen: 46 Auszahlungen, alle 46
gehen auf den Cent auf, 75 von 75 Reservierungen in Smoobu wiedergefunden.

**Airbnb ist derselbe Gedanke mit einer schwächeren Klammer.** Der Bericht ist
gleich gebaut – eine Auszahlungszeile, darunter ihre Buchungen –, aber im
Verwendungszweck der Bank steht nichts als „Airbnb". Der Referenzcode `G-…` aus
dem Bericht taucht dort nicht auf. Es bleibt der Betrag und das Datum: Airbnb
überweist am Tag der Auszahlung oder ein bis drei Tage danach. Sieben von
sieben Bewegungen ließen sich so eindeutig zuordnen. **Bleiben zwei
Auszahlungen möglich, wird nicht geraten** – dann zeigt das Tool beide.

**Die Provision kommt aus dieser Quelle**, nicht aus Smoobu – dessen Zahl hat
der Betrieb als unzuverlässig bezeichnet. Und sie besteht aus **zwei** Posten:
der Provision und der Zahlungsgebühr (`Payments Service Fee`). Wer nur die
Provision abzieht, sucht hinterher ein paar Euro je Reservierung.

Nichts hier bucht. Das Modul liest und legt ab; die Zuordnung ist B11d und
zeigt vorher, was sie tun würde.
"""
import csv
import io
import re
from datetime import date

from app import db

BOOKING = "booking"
AIRBNB = "airbnb"

# So viele Tage darf zwischen Auszahlung und Geldeingang liegen. Gemessen: 0
# bis 3 Tage. Sieben Tage lassen Luft, ohne die naechste Auszahlung
# einzufangen – Airbnb zahlt hier nie zweimal in derselben Woche denselben
# Betrag aus.
LAUFZEIT_TAGE = 7

# Das Kürzel im Verwendungszweck: `NO.vQmQNcef5aDAINyZ/ID.15049295`.
_KUERZEL = re.compile(r"\bNO\.([A-Za-z0-9]{6,})", re.I)

# Spalten des Berichts. Booking liefert sie in fester Reihenfolge, aber die
# Namen sind stabiler als die Nummern – gesucht wird über die Kopfzeile.
_SPALTEN = {
    "art": "Type/Transaction type", "schluessel": "Statement Descriptor",
    "nummer": "Reference number", "von": "Check-in date", "bis": "Check-out date",
    "wohnung_id": "Property ID", "wohnung": "Property name",
    "brutto": "Gross amount", "provision": "Commission",
    "gebuehr": "Payments Service Fee", "auszahlbar": "Payable amount",
    "betrag": "Payout amount", "datum": "Payout date",
}


def _zahl(wert):
    """'-' steht im Bericht für „leer", nicht für minus null."""
    if wert is None or wert == "" or wert == "-":
        return None
    try:
        return round(float(wert), 2)
    except (TypeError, ValueError):
        return None


def _datum(wert):
    if wert is None or wert == "" or wert == "-":
        return ""
    return str(wert)[:10]


def _text(wert):
    return "" if wert is None or wert == "-" else str(wert).strip()


def _blatt(rohdaten):
    """Das erste Tabellenblatt – oder None, wenn es keine Arbeitsmappe ist."""
    try:
        import openpyxl
    except ImportError:                                   # pragma: no cover
        return None
    if not rohdaten[:2] == b"PK":       # xlsx ist ein ZIP; alles andere spart
        return None                     # openpyxl den Versuch
    try:
        return openpyxl.load_workbook(io.BytesIO(rohdaten), read_only=True,
                                      data_only=True).worksheets[0]
    except Exception:
        return None


def art(rohdaten):
    """Welcher Bericht ist das? '' wenn keiner – dann fasst das Tool ihn nicht an."""
    ws = _blatt(rohdaten)
    if ws is not None:
        for zeile in ws.iter_rows(values_only=True):
            kopf = [_text(z) for z in zeile]
            if _SPALTEN["schluessel"] in kopf and _SPALTEN["art"] in kopf:
                return BOOKING
            break       # nur die erste Zeile; danach ist es kein Kopf mehr
        return ""
    kopf = _erste_zeile(rohdaten)
    if all(name in kopf for name in (_A["art"], _A["nummer"], _A["betrag"])):
        return AIRBNB
    return ""


def lesen(rohdaten):
    """Die Auszahlungen des Berichts, in der Reihenfolge der Datei."""
    welche = art(rohdaten)
    if welche == BOOKING:
        return _booking_lesen(rohdaten)
    if welche == AIRBNB:
        return _airbnb_lesen(rohdaten)
    return []


def _booking_lesen(rohdaten):
    """Die Auszahlungen aus dem Booking-Bericht, in der Reihenfolge der Datei.

    Eine Auszahlung ohne Reservierungen wird übergangen – sie sagt nichts, und
    es gäbe nichts zuzuordnen. Umgekehrt gilt: geht die Summe der
    Reservierungen nicht auf den Auszahlungsbetrag auf, steht das an der
    Auszahlung (`stimmt`), statt still hingenommen zu werden.
    """
    ws = _blatt(rohdaten)
    if ws is None:
        return []
    zeilen = ws.iter_rows(values_only=True)
    try:
        kopf = [_text(z) for z in next(zeilen)]
    except StopIteration:
        return []
    if _SPALTEN["schluessel"] not in kopf:
        return []
    wo = {name: kopf.index(titel) for name, titel in _SPALTEN.items()
          if titel in kopf}

    def feld(zeile, name):
        i = wo.get(name)
        return zeile[i] if i is not None and i < len(zeile) else None

    auszahlungen, nach_schluessel = [], {}
    for zeile in zeilen:
        schluessel = _text(feld(zeile, "schluessel"))
        if not schluessel:
            continue
        typ = _text(feld(zeile, "art")).strip("()").lower()
        if typ == "payout":
            a = {"portal": BOOKING, "schluessel": schluessel,
                 "datum": _datum(feld(zeile, "datum")),
                 "betrag": _zahl(feld(zeile, "betrag")) or 0.0,
                 "wohnung_id": _text(feld(zeile, "wohnung_id")),
                 "wohnung": _text(feld(zeile, "wohnung")),
                 "reservierungen": []}
            nach_schluessel[schluessel] = a
            auszahlungen.append(a)
        elif typ == "reservation":
            a = nach_schluessel.get(schluessel)
            if a is None:
                continue        # Reservierung ohne ihre Auszahlung – kommt in
                                # den echten Daten nicht vor, wäre aber kein
                                # Grund, den ganzen Bericht abzulehnen
            a["reservierungen"].append({
                "nummer": _text(feld(zeile, "nummer")),
                "von": _datum(feld(zeile, "von")),
                "bis": _datum(feld(zeile, "bis")),
                "brutto": _zahl(feld(zeile, "brutto")) or 0.0,
                "provision": _zahl(feld(zeile, "provision")) or 0.0,
                "gebuehr": _zahl(feld(zeile, "gebuehr")) or 0.0,
                "auszahlbar": _zahl(feld(zeile, "auszahlbar")) or 0.0,
                "wohnung_id": _text(feld(zeile, "wohnung_id")),
                "wohnung": _text(feld(zeile, "wohnung"))})

    raus = [a for a in auszahlungen if a["reservierungen"]]
    for a in raus:
        a["stimmt"] = _geht_auf(a)
    return raus


# --------------------------------------------------------------------- Airbnb
# Spalten des Airbnb-Berichts (deutsche Oberfläche).
_A = {"datum": "Datum", "art": "Typ", "nummer": "Bestätigungs-Code",
      "von": "Startdatum", "bis": "Enddatum", "gast": "Gast",
      "wohnung": "Inserat", "schluessel": "Referenzcode",
      "betrag": "Betrag", "auszahlung": "Ausgezahlt",
      "provision": "Servicegebühr", "gebuehr": "Gebühr für schnelle Zahlung",
      "reinigung": "Reinigungsgebühr", "brutto": "Bruttoeinkünfte",
      "steuer": "Von Airbnb abgeführte Steuer"}


def _erste_zeile(rohdaten):
    try:
        text = rohdaten.decode("utf-8-sig")
    except (AttributeError, UnicodeDecodeError):
        return []
    zeile = text.splitlines()[0] if text.splitlines() else ""
    return [s.strip().strip('"') for s in zeile.split(",")]


def _a_zahl(text):
    """Airbnb schreibt gemischt: „1562.65" und „353,42" in derselben Zeile.

    Es gibt also keine feste Schreibweise, an der man sich festhalten könnte.
    Was trägt, ist die Stellung: **das letzte Trennzeichen ist das
    Dezimaltrennzeichen**, alles davor gruppiert nur. Damit sind sowohl
    „1.234,56" als auch „1,234.56" richtig gelesen – ein Parser, der Deutsch
    annimmt, macht aus „1,916.06" die Zahl 1,92.
    """
    if text is None or str(text).strip() == "":
        return None
    s = re.sub(r"[^\d,.\-]", "", str(text).strip())
    if not s:
        return None
    letztes = max(s.rfind(","), s.rfind("."))
    if letztes >= 0:
        s = re.sub(r"[,.]", "", s[:letztes]) + "." + s[letztes + 1:]
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def _a_datum(text):
    """`04/14/2026` – Monat zuerst, wie in der amerikanischen Schreibweise."""
    teile = (text or "").strip().split("/")
    if len(teile) != 3:
        return ""
    m, t, j = teile
    return f"{j}-{m.zfill(2)}-{t.zfill(2)}" if len(j) == 4 else ""


def _airbnb_lesen(rohdaten):
    """Die Auszahlungen aus dem Airbnb-Bericht.

    Aufbau wie bei Booking: eine Zeile `Payout`, darunter die Buchungen, die
    darin stecken – zugeordnet über die Reihenfolge, denn die Buchungszeilen
    tragen den Referenzcode ihrer Auszahlung **nicht**.
    """
    try:
        text = rohdaten.decode("utf-8-sig")
    except (AttributeError, UnicodeDecodeError):
        return []
    auszahlungen, laufend = [], None
    for z in csv.DictReader(io.StringIO(text)):
        typ = (z.get(_A["art"]) or "").strip()
        if typ == "Payout":
            laufend = {
                "portal": AIRBNB, "schluessel": (z.get(_A["schluessel"]) or "").strip(),
                "datum": _a_datum(z.get(_A["datum"])),
                "betrag": _a_zahl(z.get(_A["auszahlung"])) or 0.0,
                "wohnung_id": "", "wohnung": "", "reservierungen": []}
            auszahlungen.append(laufend)
        elif typ and laufend is not None:
            # Alles andere unter einer Auszahlung ist eine Buchung – auch eine
            # „Mediations-Auszahlung". Sie ist eine Zahlung an den Betrieb wie
            # jede andere und gehört in dieselbe Auszahlung.
            brutto = _a_zahl(z.get(_A["brutto"]))
            provision = _a_zahl(z.get(_A["provision"])) or 0.0
            gebuehr = _a_zahl(z.get(_A["gebuehr"])) or 0.0
            zahlbar = _a_zahl(z.get(_A["betrag"])) or 0.0
            laufend["reservierungen"].append({
                "nummer": (z.get(_A["nummer"]) or "").strip(),
                "gast": (z.get(_A["gast"]) or "").strip(),
                "von": _a_datum(z.get(_A["von"])), "bis": _a_datum(z.get(_A["bis"])),
                # Airbnb nennt die Servicegebühr positiv; im Tool ist ein Abzug
                # negativ – wie bei Booking, sonst addierten sich die beiden
                # Portale im Verrechnungskonto gegeneinander.
                "brutto": brutto if brutto is not None else zahlbar,
                "provision": -abs(provision), "gebuehr": -abs(gebuehr),
                "auszahlbar": zahlbar,
                # Airbnb führt die Beherbergungssteuer selbst ab. Sie steht hier
                # nur zur Kenntnis – gebucht wird sie nicht, sie hat den Betrieb
                # nie erreicht.
                "steuer": _a_zahl(z.get(_A["steuer"])) or 0.0,
                "reinigung": _a_zahl(z.get(_A["reinigung"])) or 0.0,
                "wohnung_id": "", "wohnung": (z.get(_A["wohnung"]) or "").strip(),
                "art": typ})
            if not laufend["wohnung"]:
                laufend["wohnung"] = laufend["reservierungen"][-1]["wohnung"]

    raus = [a for a in auszahlungen if a["reservierungen"]]
    for a in raus:
        a["stimmt"] = _geht_auf(a)
    return raus


def _geht_auf(a):
    summe = round(sum(r["auszahlbar"] for r in a["reservierungen"]), 2)
    return abs(summe - a["betrag"]) < 0.02


def _id(a):
    return f"{a['portal']}-{a['schluessel']}"


def merken(auszahlungen):
    """Ablegen, was noch nicht da ist. Gibt (neu, schon bekannt) zurück.

    Der Betrieb lädt den Bericht immer als „01.01. bis heute" – jede zweite
    Datei enthält alles aus der ersten. Der Schlüssel ist deshalb das Kürzel
    der Auszahlung, nicht die Datei: derselbe Bericht zweimal ändert nichts.

    Enthält der neue Bericht zu einer bekannten Auszahlung **mehr**
    Reservierungen, werden sie ergänzt. Das ist der Fall, in dem der erste
    Export mitten in einer Abrechnungswoche endete.
    """
    neu = doppelt = 0
    with db.transaktion():
        for a in auszahlungen:
            vorher = db.holen("portalauszahlungen", _id(a))
            if vorher is None:
                db.anlegen("portalauszahlungen", dict(a), _id(a))
                neu += 1
                continue
            doppelt += 1
            bekannt = {r.get("nummer") for r in vorher.get("reservierungen", [])}
            fehlend = [r for r in a["reservierungen"] if r.get("nummer") not in bekannt]
            if fehlend:
                vorher["reservierungen"] = vorher.get("reservierungen", []) + fehlend
                vorher["stimmt"] = _geht_auf(vorher)
                db.speichern("portalauszahlungen", _id(a), vorher)
    return neu, doppelt


def auszahlungen(portal=""):
    """Alle abgelegten Auszahlungen, älteste zuerst."""
    alle = db.alle("portalauszahlungen")
    if portal:
        alle = [a for a in alle if a.get("portal") == portal]
    return sorted(alle, key=lambda a: (a.get("datum", ""), a.get("schluessel", "")))


def kuerzel(bewegung):
    """Das Auszahlungskürzel im Verwendungszweck – '' wenn keins."""
    treffer = _KUERZEL.search((bewegung.get("text") or "") + " "
                              + (bewegung.get("gegenpartei") or ""))
    return treffer.group(1) if treffer else ""


def _tage(a, b):
    """Tage zwischen zwei Datumsangaben – None, wenn eine unlesbar ist."""
    try:
        x = date(*[int(t) for t in a.split("-")])
        y = date(*[int(t) for t in b.split("-")])
    except (ValueError, TypeError, AttributeError):
        return None
    return (x - y).days


def kandidaten(bewegung):
    """Welche Auszahlungen zu dieser Bankbewegung passen könnten.

    Zwei Wege, weil die Portale zwei verschiedene Klammern liefern:

    * **Booking** nennt sein Kürzel im Verwendungszweck. Das ist eindeutig; wo
      es fehlt, gehört die Bewegung eben nicht zu einer Auszahlung.
    * **Airbnb** nennt nichts. Dort passt eine Auszahlung, wenn der Betrag
      stimmt und der Geldeingang am Tag der Auszahlung oder bis zu einer Woche
      danach liegt.

    Die Liste hat in aller Regel null oder einen Eintrag. Hat sie zwei, ist der
    Fall mehrdeutig – und das ist eine Antwort, kein Fehler.
    """
    betrag = round(bewegung.get("betrag", 0) or 0, 2)
    if betrag <= 0:
        return []
    k = kuerzel(bewegung)
    if k:
        treffer = db.holen("portalauszahlungen", f"{BOOKING}-{k}")
        return [treffer] if treffer else []
    from app import verrechnung
    if verrechnung.plattform_von(bewegung) != AIRBNB:
        return []
    raus = []
    for a in auszahlungen(AIRBNB):
        abstand = _tage(bewegung.get("datum") or "", a.get("datum") or "")
        # Auf den Cent genau: die Betraege sind Cent-Zahlen, keine Schaetzungen.
        # Eine Toleranz von einem Cent liesse zwei verschiedene Auszahlungen
        # als denselben Treffer durchgehen.
        if abstand is not None and 0 <= abstand <= LAUFZEIT_TAGE \
                and round(a.get("betrag", 0) - betrag, 2) == 0:
            raus.append(a)
    return raus


def zu_bewegung(bewegung):
    """Die Auszahlung zu dieser Bankbewegung – oder None.

    None heißt „keine gefunden" **oder** „mehr als eine möglich". Geraten wird
    nicht: eine falsch zugeordnete Auszahlung erzeugt eine Rechnungszuordnung,
    die niemand mehr in Frage stellt.
    """
    treffer = kandidaten(bewegung)
    return treffer[0] if len(treffer) == 1 else None


def einlesen(rohdaten):
    """Ist das ein Auszahlungsbericht? Dann einlesen und ablegen, sonst None.

    Damit wandert der Bericht durch **dieselbe** Auswahl wie die Kontoauszüge.
    Der Betrieb hat die vielen getrennten Schritte ausdrücklich als
    fehleranfällig benannt; eine eigene Hochladestelle je Dateiart wäre einer
    mehr. Was die Datei ist, erkennt das Tool an ihrem Inhalt.
    """
    welche = art(rohdaten)
    if not welche:
        return None
    gelesen = lesen(rohdaten)
    neu, doppelt = merken(gelesen)
    tage = sorted(a["datum"] for a in gelesen if a.get("datum"))
    from app import verrechnung
    return {"art": "portal", "portal": welche, "konto": verrechnung.name_von(welche),
            "neu": neu, "doppelt": doppelt,
            "auszahlungen": len(gelesen),
            "reservierungen": sum(len(a["reservierungen"]) for a in gelesen),
            "schief": [a["schluessel"] for a in gelesen if not a["stimmt"]],
            "zeitraum": (tage[0], tage[-1]) if tage else ("", "")}
